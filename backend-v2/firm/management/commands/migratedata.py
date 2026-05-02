"""
Migrate data from Excel files in the ``migrationdata/`` directory.

Creates two firms (Anmol Corner, Anmol Enterprise) owned by a single user,
then imports:
  - Products   from files whose name contains "item" or "product"  (item master)
  - Customers  from files whose name contains "customer" or "retailer"

Usage:
    python manage.py migratedata
    python manage.py migratedata --skip-firms   # only import XLS, skip firm/user creation
    python manage.py migratedata --refresh-products  # re-apply rates/MRP from XLS to existing rows (by item name)
    python manage.py migratedata --skip-firms --refresh-customer-discounts  # set default_discount_percent from Dis %% etc.
    python manage.py migratedata --skip-firms --import-stock  # sync stock from *anmol corner.xls* / *anmol enterprise.xls*
"""

import os
import re
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction, connection
from django.db.utils import OperationalError
from django.utils.text import slugify

from django.db.models import Sum

from accounts.models import User, FirmUsers
from firm.choices import StockLedgerEntryType, StockManualReason
from firm.models import Firm, Product, Customer, ProductBatch, StockLedgerEntry


OWNER_EMAIL = "admin@mail.com"
OWNER_PASSWORD = "Admin@123"
OWNER_FULL_NAME = "Super Admin"
OWNER_PHONE = "9000000000"

FIRMS = [
    {"name": "Anmol Corner", "code": "ANMOLCNR"},
    {"name": "Anmol Enterprise", "code": "ANMOLENT"},
]

MIGRATION_DIR = Path(settings.BASE_DIR) / "migrationdata"


# -- helpers (same logic as firm.apis) --


def _row_get(row: dict, *keys):
    for k in keys:
        if k in row and row[k] is not None and str(row[k]).strip() != "":
            return row[k]
    return None


def _dec(val, default=Decimal("0")):
    if val is None or str(val).strip() == "":
        return default
    try:
        return Decimal(str(val).replace(",", "").strip())
    except Exception:
        return default


def _parse_active(val):
    if val is None or str(val).strip() == "":
        return True
    return str(val).strip().upper() not in ("NO", "FALSE", "0", "N", "INACTIVE")


def _find_header_row(raw_rows: list[list], min_filled=3) -> int:
    """Return the index of the first row where at least *min_filled* cells look
    like column headers (non-empty strings).  Skips title / banner rows."""
    for idx, row in enumerate(raw_rows):
        non_empty_str = sum(
            1 for v in row
            if v is not None and isinstance(v, str) and v.strip()
        )
        if non_empty_str >= min_filled:
            return idx
    return 0


def _read_xls_rows(filepath: str) -> list[dict]:
    """Read rows from .xls (legacy binary) file using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(filepath)
    sheet = wb.sheet_by_index(0)
    if sheet.nrows < 2:
        return []

    raw = []
    for r in range(sheet.nrows):
        raw.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])

    hdr_idx = _find_header_row(raw)
    headers = [str(raw[hdr_idx][c]).strip().lower() for c in range(sheet.ncols)]

    rows = []
    for r in range(hdr_idx + 1, len(raw)):
        vals = raw[r]
        if not any(str(v).strip() for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


def _read_xlsx_rows(filepath: str) -> list[dict]:
    """Read rows from .xlsx file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return []

    hdr_idx = _find_header_row(all_rows)
    headers = [str(h).strip().lower() if h else "" for h in all_rows[hdr_idx]]

    rows = []
    for row in all_rows[hdr_idx + 1:]:
        if not any(row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _read_excel(filepath: str) -> list[dict]:
    if filepath.lower().endswith(".xls"):
        return _read_xls_rows(filepath)
    return _read_xlsx_rows(filepath)


def _read_stock_excel(filepath: str) -> list[dict]:
    """
    Stock workbooks often start with vendor banner rows; headers are like
    'Item Name' / 'Stock Quantity'. Generic _find_header_row would pick the
    wrong row (e.g. Vendor Name / GST / Address).
    """
    if filepath.lower().endswith(".xls"):
        import xlrd

        wb = xlrd.open_workbook(filepath)
        sh = wb.sheet_by_index(0)
        raw = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        sh = wb.active
        raw = list(sh.iter_rows(values_only=True))

    hdr_idx = None
    for idx, row in enumerate(raw):
        headers_low = [str(v).strip().lower() if v is not None else "" for v in row]
        nonempty = [h for h in headers_low if h]
        blob = " ".join(nonempty)
        has_item = "item" in blob and "name" in blob
        has_qty_col = (
            "stock quantity" in blob
            or ("stock" in blob and "quantity" in blob)
            or blob.replace(" ", "").find("stockqty") >= 0
        )
        if has_item and has_qty_col:
            hdr_idx = idx
            break

    if hdr_idx is None:
        return _read_excel(filepath)

    headers = [
        str(raw[hdr_idx][c]).strip().lower() if c < len(raw[hdr_idx]) else ""
        for c in range(len(raw[hdr_idx]))
    ]
    rows = []
    for r in range(hdr_idx + 1, len(raw)):
        vals = raw[r]
        if vals is None:
            continue
        if not any(str(v).strip() for v in vals if v is not None):
            continue
        row_dict = dict(zip(headers, vals))
        rows.append(row_dict)
    return rows


# -- importers --


def _product_fields_from_row(row: dict) -> dict:
    """Map one Excel row to Product field values (shared by create + refresh)."""
    def _max(s: str | None, n: int) -> str | None:
        if not s:
            return None
        s2 = str(s).strip()
        return s2[:n] if len(s2) > n else s2

    return {
        "product_code": _max(
            str(
            _row_get(row, "product_code", "productcode", "code", "item code", "itemcode", "sr no") or ""
            ).strip()
            or None,
            64,
        ),
        "description": (str(_row_get(row, "description") or "").strip() or None),
        "category": _max((str(_row_get(row, "category") or "").strip() or None), 500),
        "hsn_code": _max(
            (str(_row_get(row, "hsnno", "hsn_no", "hsn_code", "hsn code", "hsn") or "").strip() or None),
            50,
        ),
        "gst_percent": _dec(_row_get(row, "gstper", "gst_percent", "gst %", "gst")),
        "liters": _dec(_row_get(row, "liters", "ltr")),
        "pack": _dec(_row_get(row, "pack")),
        "mrp": _dec(_row_get(row, "mrp")),
        "purchase_rate": _dec(_row_get(row, "prate", "purchase_rate", "purchase rate")),
        "purchase_rate_per_unit": _dec(
            _row_get(row, "prateunit", "prateur", "purchase_rate_per_unit")
        ),
        "sale_rate": _dec(_row_get(row, "srate", "sale_rate", "sale rate", "s_rate")),
        # Excel exports use header "Rate/unit" -> key "rate/unit" (not "rate/uni")
        "rate_per_unit": _dec(
            _row_get(
                row,
                "rate/unit",
                "rate/uni",
                "rate_uni",
                "rate per unit",
                "rate_per_unit",
            )
        ),
        "product_discount": _dec(_row_get(row, "product_discount", "discount", "disc")),
        "is_active": _parse_active(_row_get(row, "active", "is_active")),
    }


def import_products(firm: Firm, rows: list[dict], stdout) -> int:
    created = 0
    skipped = 0
    existing_names = {
        (n or "").strip().lower()
        for n in Product.objects.filter(firm=firm).values_list("name", flat=True)
    }
    for i, row in enumerate(rows):
        name_raw = _row_get(row, "itemname", "item name", "name", "product name", "item_name")
        name = str(name_raw).strip() if name_raw else ""
        if not name:
            stdout.write(f"  [products] Row {i+2}: skipped – no name")
            skipped += 1
            continue

        key = name.strip().lower()
        if key in existing_names:
            stdout.write(f"  [products] Row {i+2}: '{name}' already exists, skipping")
            skipped += 1
            continue

        pf = _product_fields_from_row(row)
        Product.objects.create(
            firm=firm,
            name=name,
            **pf,
        )
        existing_names.add(key)
        created += 1

    stdout.write(f"  [products] {created} created, {skipped} skipped for {firm.name}")
    return created


def refresh_products_from_rows(firm: Firm, rows: list[dict], stdout) -> tuple[int, int]:
    """Update existing products matched by name (case-insensitive) from spreadsheet."""
    updated = 0
    missing = 0
    for row in rows:
        name_raw = _row_get(row, "itemname", "item name", "name", "product name", "item_name")
        name = str(name_raw).strip() if name_raw else ""
        if not name:
            continue
        qs = Product.objects.filter(firm=firm, name__iexact=name)
        if not qs.exists():
            missing += 1
            continue
        pf = _product_fields_from_row(row)
        qs.update(**pf)
        updated += 1
    stdout.write(f"  [products refresh] {updated} updated, {missing} spreadsheet rows had no DB match for {firm.name}")
    return updated, missing


def _clean_str(val) -> str:
    """Convert cell value to a clean string. Handles floats like 9028680353.0 -> '9028680353'."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def _excel_date_to_date(val):
    """Convert an Excel serial date number to a timezone-aware datetime, or None."""
    from datetime import datetime, timedelta
    from django.utils import timezone as tz
    if val is None:
        return None
    try:
        serial = float(val)
        if serial <= 0:
            return None
        naive = datetime(1899, 12, 30) + timedelta(days=serial)
        return tz.make_aware(naive)
    except (ValueError, TypeError):
        return None


def _customer_discount_from_row(row: dict):
    """Customer masters use headers like 'Dis %' → dict key 'dis %' after normalize."""
    return _dec(
        _row_get(
            row,
            "default_discount_percent",
            "default_discount",
            "retailer_discount",
            "discount",
            "dis %",
            "disc %",
            "dis%",
            "disc%",
        )
    )


def refresh_customer_discounts_from_rows(firm: Firm, rows: list[dict], stdout) -> tuple[int, int]:
    """Set default_discount_percent from spreadsheet rows matched by business name."""
    updated = 0
    missing = 0
    for row in rows:
        biz_name = _clean_str(
            _row_get(
                row,
                "name",
                "business_name",
                "business name",
                "businessname",
                "shop name",
                "shopname",
                "shop_name",
            )
        )
        if not biz_name:
            continue
        new_disc = _customer_discount_from_row(row)
        n = Customer.objects.filter(firm=firm, business_name__iexact=biz_name).update(
            default_discount_percent=new_disc
        )
        if n:
            updated += n
        else:
            missing += 1
    stdout.write(
        f"  [customer discounts] {updated} rows updated, "
        f"{missing} spreadsheet rows had no DB customer for {firm.name}"
    )
    return updated, missing


def import_customers(firm: Firm, rows: list[dict], stdout) -> int:
    created = 0
    skipped = 0
    existing_names = {
        (n or "").strip().lower()
        for n in Customer.objects.filter(firm=firm).values_list("business_name", flat=True)
    }
    for i, row in enumerate(rows):
        biz_name = _clean_str(
            _row_get(row, "name", "business_name", "business name", "businessname", "shop name", "shopname", "shop_name")
        )
        if not biz_name:
            skipped += 1
            continue

        biz_key = biz_name.strip().lower()
        if biz_key in existing_names:
            skipped += 1
            continue

        owner_name = _clean_str(_row_get(row, "owner_name", "owner name", "ownername", "owner", "contact_person", "contact person"))
        mobile_raw = _clean_str(_row_get(row, "mobileno", "mobile no", "mobile", "whatsapp_number", "whatsapp", "phone", "contact_number", "contact number"))
        address = _clean_str(_row_get(row, "address", "business_address", "business address"))
        email = _clean_str(_row_get(row, "email", "email_id", "email id"))
        area = _clean_str(_row_get(row, "area"))

        gst_raw = _clean_str(_row_get(row, "gstin", "gst_number", "gst number", "gst"))
        gst = (gst_raw[:50] if gst_raw and gst_raw != "0" else None)

        fssai_raw = _clean_str(_row_get(row, "fssai", "fssai_number", "fssai number"))
        fssai = (fssai_raw[:50] if fssai_raw and fssai_raw != "0" else None)

        fssai_expiry = _excel_date_to_date(_row_get(row, "exp date", "expdate", "fssai_expiry", "expiry"))

        ref_id = _clean_str(_row_get(row, "id", "reference_code", "ref_code", "ref code", "code"))
        ref_code = f"R{ref_id}" if ref_id else None

        full_address = address
        if area:
            full_address = f"{address}, {area}" if address else area

        ctype_raw = _clean_str(_row_get(row, "invtype", "customer_type", "type", "customer type")).upper()
        customer_type = "DISTRIBUTOR"
        if ctype_raw in ("SUPER_SELLER", "SUPERSELLER", "SS"):
            customer_type = "SUPER_SELLER"

        cust_slug = slugify(f"{biz_name}-{firm.code}-{uuid4().hex[:6]}")[:50]

        default_disc = _customer_discount_from_row(row)

        payload = dict(
            firm=firm,
            slug=cust_slug,
            customer_type=customer_type,
            business_name=biz_name,
            owner_name=owner_name or biz_name,
            whatsapp_number=mobile_raw or "0000000000",
            contact_number=mobile_raw or "0000000000",
            email=email or f"{slugify(biz_name)}@placeholder.com",
            business_address=full_address or "N/A",
            gst_number=gst,
            fssai_number=fssai,
            fssai_expiry=fssai_expiry,
            reference_code=ref_code,
            default_discount_percent=default_disc,
            is_active=True,
        )
        try:
            Customer.objects.create(**payload)
        except OperationalError:
            # Resiliency for flaky remote DB connections: force reconnect and retry once.
            try:
                connection.close()
            except Exception:
                pass
            Customer.objects.create(**payload)
        existing_names.add(biz_key)
        created += 1

    stdout.write(f"  [customers] {created} created, {skipped} skipped for {firm.name}")
    return created


def _int_stock(val) -> int:
    if val is None or str(val).strip() == "":
        return 0
    try:
        return int(round(float(str(val).replace(",", "").strip())))
    except Exception:
        return 0


def _workbook_looks_like_stock_sheet(rows: list[dict]) -> bool:
    """True if headers include a stock quantity column (not item-master only)."""
    if not rows:
        return False
    keys = {(str(k or "").strip().lower()) for k in rows[0].keys()}
    if keys & {"stock", "stockpcs", "stock pcs", "closing stock", "stock quantity"}:
        return True
    return any(
        ("stock" in k and "quantity" in k) or k.endswith(" stock") or k.startswith("stock ")
        for k in keys
        if k
    )


def _sync_product_stock_absolute(firm: Firm, product: Product, target_qty: int) -> str:
    """
    Make sum(ProductBatch.quantity) == target_qty by adjusting the single no-expiry batch.
    Records StockLedgerEntry OPENING_BALANCE for audit.
    """
    agg = ProductBatch.objects.filter(product=product).aggregate(s=Sum("quantity"))
    current = int(agg["s"] or 0)
    delta = int(target_qty) - current
    if delta == 0:
        return "unchanged"

    batch = ProductBatch.objects.filter(product=product, expiry_date__isnull=True).first()
    if batch is None:
        batch = ProductBatch.objects.create(product=product, expiry_date=None, quantity=0)

    batch.quantity += delta
    batch.save(update_fields=["quantity", "updated_on"])
    StockLedgerEntry.objects.create(
        firm=firm,
        product=product,
        product_batch=batch,
        quantity_delta=delta,
        entry_type=StockLedgerEntryType.MANUAL,
        manual_reason=StockManualReason.OPENING_BALANCE,
        note="migratedata: Excel stock sheet sync",
        created_by=None,
    )
    return "updated"


def _strip_leading_erp_item_prefix(name: str) -> str | None:
    """Sheets like '10693/Plain Pista ...' often omit the numeric prefix in product master."""
    s = name.strip()
    if "/" not in s[:24]:
        return None
    tail = s.split("/", 1)[1].strip()
    return tail or None


def _norm_stock_key(name: str) -> str:
    """Lowercase + collapse whitespace for stable alias / placeholder keys."""
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _stock_import_name_aliases(firm_code: str) -> dict[str, str]:
    """
    Map normalized Excel item cell -> exact Product.name when ERP export text
    differs from the catalog (spacing, price suffix, ml variant, etc.).
    """
    if firm_code != "ANMOLCNR":
        return {}
    nk = _norm_stock_key
    return {
        nk("10693/Plain Pista BP 5000ml 1PR570 C6-Creambell"): "Plain Pista BP 5000ml 1P R545",
        nk("22175/Royal Rajwadi 40ml 24P R10C9N-Creambell"): "Royal Rajwadi Kulfi 50ml 24P R10 [newR9]",
        nk("13070/Butter Crunch 70ml 18P R25C9-Creambell"): "Butter Crunch Choc 70ml 18P R25 (New R20)",
        nk("14016/Cassatta Chops 150ml 10P R70C7-Creambell"): "Cassatta Chops 150ML 10P R50",
        nk("35069/Vanilla Jumbo Pack 2000ml 1P R285C15-Creambell"): "Vanilla Jumbo Pack 2000ML 1P R200",
        nk("26139/Star Bon Vanilla 15ml 60P R5C9-Creambell"): "Star Bon Vanilla 15ml 60P R4.5",
        nk("13104/Choco Vanilla Choc 65ml 24P R20C9-Creambell"): "Choco Vanilla Choc 65ml 24P R20 (New R18)",
        nk("22125/ Mawa Malai Kulfi 40ml 24PR10 C9-Creambell"): "Mawa Malai Kulfi 40ml 24P R9",
        nk("22125/Mawa Malai Kulfi 40ml 24PR10 C9-Creambell"): "Mawa Malai Kulfi 40ml 24P R9",
        nk("18350/Butter Scotch Cup 85ml 12PR20C6-Creambell"): "Butter Scotch Cup 85ML 16P R18",
        nk("16325/D.Chocolate Cone 105ml20PR40C6-Creambell"): "Chocolate Cone 105ml 20P R35",
    }


def _anmol_cnr_stock_placeholder_norm_keys() -> frozenset[str]:
    """Sheet rows with no existing catalog match: create a minimal Product using the sheet label."""
    nk = _norm_stock_key
    return frozenset(
        {
            nk("16338/Cookies & Cream Disc Cone 105ml 10PR50C9-Creambell"),
            nk("16339/Nutty Chocolate Disccone 105ml 10PR60C9-Creambell"),
            nk("18348/Belgian Cookie Cup 110ml8PR50C9-Creambell"),
            nk("18367/Royal Dry Fruit Cup 110ml 8PR40C9-Creambell"),
            nk("16326/Crunchy Butterscotch Cone 105ml20PR30C6-Creambell"),
        }
    )


def _resolve_product_for_stock_import(firm: Firm, sheet_name: str) -> Product | None:
    name = (sheet_name or "").strip()
    if not name:
        return None

    qs = Product.objects.filter(firm=firm, name__iexact=name)
    if qs.exists():
        return qs.first()
    alt = _strip_leading_erp_item_prefix(name)
    if alt:
        qs = Product.objects.filter(firm=firm, name__iexact=alt)
        if qs.exists():
            return qs.first()

    db_name = _stock_import_name_aliases(firm.code).get(_norm_stock_key(name))
    if db_name:
        qs = Product.objects.filter(firm=firm, name__iexact=db_name)
        if qs.exists():
            return qs.first()

    if firm.code == "ANMOLCNR" and _norm_stock_key(name) in _anmol_cnr_stock_placeholder_norm_keys():
        nm = name[:255]
        existing = Product.objects.filter(firm=firm, name__iexact=nm).first()
        if existing:
            return existing
        return Product.objects.create(firm=firm, name=nm)

    return None


def import_stock_from_rows(firm: Firm, rows: list[dict], stdout) -> tuple[int, int, int]:
    """
    Match rows by item name (case-insensitive); read Stock / StockPcs / etc.
    Returns (updated, unchanged, missing_product).
    """
    updated = unchanged = missing = 0
    missing_samples: list[str] = []

    for row in rows:
        name_raw = _row_get(row, "itemname", "item name", "name", "product name", "item_name")
        name = str(name_raw).strip() if name_raw else ""
        if not name:
            continue

        stock_raw = _row_get(
            row,
            "stock quantity",
            "stockquantity",
            "stock qty",
            "stockqty",
            "stock",
            "stockpcs",
            "stock pcs",
            "closing stock",
            "qty",
            "balance",
        )
        target = _int_stock(stock_raw)

        product = _resolve_product_for_stock_import(firm, name)
        if product is None:
            missing += 1
            if len(missing_samples) < 15:
                missing_samples.append(name[:120])
            continue

        res = _sync_product_stock_absolute(firm, product, target)
        if res == "updated":
            updated += 1
        else:
            unchanged += 1

    stdout.write(
        f"  [stock] {updated} updated, {unchanged} already correct, "
        f"{missing} spreadsheet rows had no DB product for {firm.name}"
    )
    if missing_samples:
        stdout.write(f"  [stock] sample unmatched names: {missing_samples[:5]}")
    return updated, unchanged, missing


# -- management command --


class Command(BaseCommand):
    help = "Create Anmol Corner / Anmol Enterprise firms and import product + customer data from migrationdata/*.xls(x)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--skip-firms",
            action="store_true",
            help="Skip firm/user creation (only run XLS imports)",
        )
        parser.add_argument(
            "--refresh-products",
            action="store_true",
            help="Re-read item-master XLS and update existing products (rates, MRP, etc.) by item name; does not create rows",
        )
        parser.add_argument(
            "--import-stock",
            action="store_true",
            help="Sync on-hand stock from XLS files that match each firm name and contain Stock / StockPcs columns",
        )
        parser.add_argument(
            "--refresh-customer-discounts",
            action="store_true",
            help=(
                "Re-read customer/retailer XLS and set Customer.default_discount_percent "
                "(matches 'Dis %%', discount columns) by shop name"
            ),
        )

    def handle(self, *args, **options):
        if options["import_stock"]:
            firms = []
            for fd in FIRMS:
                try:
                    firms.append(Firm.objects.get(code=fd["code"]))
                except Firm.DoesNotExist:
                    self.stderr.write(
                        self.style.ERROR(
                            f"Firm {fd['code']} not found – create firms first (run migratedata without --import-stock only)"
                        )
                    )
                    return
            self._import_stock_from_files(firms)
            self.stdout.write(self.style.SUCCESS("\nStock import complete!"))
            return

        if options["refresh_products"]:
            firms = []
            for fd in FIRMS:
                try:
                    firms.append(Firm.objects.get(code=fd["code"]))
                except Firm.DoesNotExist:
                    self.stderr.write(
                        self.style.ERROR(f"Firm {fd['code']} not found – run migratedata without --refresh-products first")
                    )
                    return
            self._refresh_products_from_files(firms)
            self.stdout.write(self.style.SUCCESS("\nProduct refresh complete!"))
            return

        if options["refresh_customer_discounts"]:
            firms = []
            for fd in FIRMS:
                try:
                    firms.append(Firm.objects.get(code=fd["code"]))
                except Firm.DoesNotExist:
                    self.stderr.write(
                        self.style.ERROR(
                            f"Firm {fd['code']} not found – create firms first (run `python manage.py migratedata`)"
                        )
                    )
                    return
            self._refresh_customer_discounts_from_files(firms)
            self.stdout.write(self.style.SUCCESS("\nCustomer discount refresh complete!"))
            return

        firms = []

        if not options["skip_firms"]:
            owner = self._ensure_owner()
            firms = self._ensure_firms(owner)
        else:
            for fd in FIRMS:
                try:
                    firms.append(Firm.objects.get(code=fd["code"]))
                except Firm.DoesNotExist:
                    self.stderr.write(self.style.ERROR(f"Firm {fd['code']} not found – run without --skip-firms first"))
                    return

        self._import_from_files(firms)
        self.stdout.write(self.style.SUCCESS("\nMigration complete!"))

    # -- firm / user setup --

    def _ensure_owner(self) -> User:
        # Create a *global* admin user (ADMIN) as requested.
        user = User.objects.filter(email=OWNER_EMAIL).first()
        if user is None:
            user = User.objects.create_superuser(
                phone=OWNER_PHONE,
                email=OWNER_EMAIL,
                full_name=OWNER_FULL_NAME,
                password=OWNER_PASSWORD,
            )
            self.stdout.write(self.style.SUCCESS(f"  Created admin user: {OWNER_EMAIL}"))
            return user

        # Ensure existing user is elevated to ADMIN and has the requested password.
        user.is_admin = True
        user.is_active = True
        user.user_type = "ADMIN"
        user.full_name = user.full_name or OWNER_FULL_NAME
        user.phone = user.phone or OWNER_PHONE
        user.set_password(OWNER_PASSWORD)
        user.save()
        self.stdout.write(f"  Admin user already exists; ensured ADMIN + reset password: {OWNER_EMAIL}")
        return user

    def _ensure_firms(self, owner: User) -> list[Firm]:
        firms = []
        for fd in FIRMS:
            firm, created = Firm.objects.get_or_create(
                code=fd["code"],
                defaults={"name": fd["name"]},
            )
            tag = "Created" if created else "Exists"
            self.stdout.write(f"  {tag} firm: {firm.name} ({firm.slug})")

            FirmUsers.objects.get_or_create(
                user=owner,
                firm=firm,
                defaults={"role": "FIRM_ADMIN"},
            )
            firms.append(firm)
        return firms

    # -- XLS import --

    @staticmethod
    def _file_matches_firm(filepath: Path, firm_name_lower: str) -> bool:
        """Check if a file name corresponds to a firm. Handles partial matches
        like 'anmol enterprise' matching 'Anmol Enterprises Item master'."""
        stem = filepath.stem.lower()
        return all(w in stem for w in firm_name_lower.split())

    def _import_from_files(self, firms: list[Firm]):
        if not MIGRATION_DIR.exists():
            self.stdout.write(self.style.WARNING(
                f"  migrationdata/ directory not found at {MIGRATION_DIR} – skipping file imports"
            ))
            return

        xls_files = [
            f for f in MIGRATION_DIR.iterdir()
            if f.suffix.lower() in (".xls", ".xlsx") and not f.name.startswith("~")
        ]

        if not xls_files:
            self.stdout.write(self.style.WARNING("  No .xls/.xlsx files found in migrationdata/"))
            return

        product_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("item", "product"))]
        customer_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("customer", "retailer"))]
        unmatched = [f for f in xls_files if f not in product_files and f not in customer_files]

        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"  Unmatched files (not imported): {[f.name for f in unmatched]}"
            ))

        for firm in firms:
            self.stdout.write(f"\n  -- Importing for {firm.name} --")
            firm_key = firm.name.lower()

            firm_product_files = [f for f in product_files if self._file_matches_firm(f, firm_key)]
            firm_customer_files = [f for f in customer_files if self._file_matches_firm(f, firm_key)]

            if not firm_product_files and not firm_customer_files:
                self.stdout.write(f"  No matching files found for {firm.name}")
                continue

            for pf in firm_product_files:
                self.stdout.write(f"  Reading products from {pf.name} ...")
                rows = _read_excel(str(pf))
                import_products(firm, rows, self.stdout)

            for cf in firm_customer_files:
                self.stdout.write(f"  Reading customers from {cf.name} ...")
                rows = _read_excel(str(cf))
                import_customers(firm, rows, self.stdout)

    def _refresh_customer_discounts_from_files(self, firms: list[Firm]):
        if not MIGRATION_DIR.exists():
            self.stdout.write(self.style.WARNING(f"  migrationdata/ not found at {MIGRATION_DIR}"))
            return
        xls_files = [
            f for f in MIGRATION_DIR.iterdir()
            if f.suffix.lower() in (".xls", ".xlsx") and not f.name.startswith("~")
        ]
        customer_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("customer", "retailer"))]
        if not customer_files:
            self.stdout.write(self.style.WARNING("  No customer/retailer XLS files in migrationdata/"))
            return
        for firm in firms:
            self.stdout.write(f"\n  -- Refresh customer discounts for {firm.name} --")
            firm_key = firm.name.lower()
            for cf in [f for f in customer_files if self._file_matches_firm(f, firm_key)]:
                self.stdout.write(f"  Updating from {cf.name} ...")
                rows = _read_excel(str(cf))
                refresh_customer_discounts_from_rows(firm, rows, self.stdout)

    def _refresh_products_from_files(self, firms: list[Firm]):
        if not MIGRATION_DIR.exists():
            self.stdout.write(self.style.WARNING(f"  migrationdata/ not found at {MIGRATION_DIR}"))
            return
        xls_files = [
            f for f in MIGRATION_DIR.iterdir()
            if f.suffix.lower() in (".xls", ".xlsx") and not f.name.startswith("~")
        ]
        product_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("item", "product"))]
        if not product_files:
            self.stdout.write(self.style.WARNING("  No item/product XLS files in migrationdata/"))
            return
        for firm in firms:
            self.stdout.write(f"\n  -- Refresh products for {firm.name} --")
            firm_key = firm.name.lower()
            for pf in [f for f in product_files if self._file_matches_firm(f, firm_key)]:
                self.stdout.write(f"  Updating from {pf.name} ...")
                with transaction.atomic():
                    rows = _read_excel(str(pf))
                    refresh_products_from_rows(firm, rows, self.stdout)

    def _import_stock_from_files(self, firms: list[Firm]):
        if not MIGRATION_DIR.exists():
            self.stdout.write(self.style.WARNING(f"  migrationdata/ not found at {MIGRATION_DIR}"))
            return

        xls_files = [
            f for f in MIGRATION_DIR.iterdir()
            if f.suffix.lower() in (".xls", ".xlsx") and not f.name.startswith("~")
        ]
        if not xls_files:
            self.stdout.write(self.style.WARNING("  No .xls/.xlsx files in migrationdata/"))
            return

        product_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("item", "product"))]
        customer_files = [f for f in xls_files if any(k in f.stem.lower() for k in ("customer", "retailer"))]

        for firm in firms:
            self.stdout.write(f"\n  -- Stock import for {firm.name} --")
            firm_key = firm.name.lower()
            for xf in xls_files:
                if xf in product_files or xf in customer_files:
                    continue
                # Avoid accidentally parsing large unrelated masters named "*corner*" etc.
                if "stock" not in xf.stem.lower():
                    continue
                if not self._file_matches_firm(xf, firm_key):
                    continue
                rows = _read_stock_excel(str(xf))
                if not _workbook_looks_like_stock_sheet(rows):
                    self.stdout.write(f"  Skip {xf.name} (no Stock / StockPcs columns)")
                    continue
                self.stdout.write(f"  Applying stock from {xf.name} ...")
                import_stock_from_rows(firm, rows, self.stdout)
