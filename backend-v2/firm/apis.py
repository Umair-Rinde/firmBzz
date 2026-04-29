from decimal import Decimal

from .models import (
    Firm,
    Product,
    VendorOrder,
    VendorOrderItem,
    Vendor,
    Customer,
    Invoice,
    InvoiceItem,
    ProductBatch,
    RetailerOrder,
    Payment,
    StockLedgerEntry,
)
from accounts.models import FirmUsers
from .serializers import (
    FirmSerializer,
    ProductSerializer,
    VendorOrderSerializer,
    VendorOrderCreateSerializer,
    FirmUserSerializer,
    FirmUserCreateSerializer,
    FirmUserUpdateSerializer,
    VendorSerializer,
    CustomerSerializer,
    InvoiceSerializer,
    InvoiceFromRetailerOrdersSerializer,
    InvoiceUpdateSerializer,
    build_lines_from_line_overrides,
    apply_invoice_lines_to_invoice,
    reverse_invoice_line_allocations,
    RetailerOrderSerializer,
    RetailerOrderCreateSerializer,
    PaymentSerializer,
    PaymentCreateSerializer,
    StockManualAdjustSerializer,
)
from .pricing import effective_unit_rate, allocate_batches_fefo, line_total_inclusive
from .choices import StockLedgerEntryType
from rest_framework.exceptions import ValidationError as DRFValidationError
from portal.base import BaseResponse
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, F, Sum


def _get_pg_limit_q(params):
    try:
        pg = int(params.get("pg") or params.get("page") or 0)
    except:
        pg = 0
    try:
        limit = int(params.get("limit") or 20)
    except:
        limit = 20
    if pg < 0:
        pg = 0
    if limit <= 0:
        limit = 20
    q = (params.get("q") or params.get("search") or "").strip()
    return pg, limit, q


def _drf_validation_detail_message(exc: DRFValidationError) -> str:
    """Turn DRF ValidationError.detail (list/dict/str) into one message for BaseResponse."""
    detail = getattr(exc, "detail", None)
    if detail is None:
        return "Request failed validation."
    if isinstance(detail, list):
        return str(detail[0]) if detail else "Request failed validation."
    if isinstance(detail, dict):
        parts = []
        for val in detail.values():
            if isinstance(val, list):
                parts.extend(str(x) for x in val)
            else:
                parts.append(str(val))
        return "; ".join(parts) if parts else str(detail)
    return str(detail)


def _build_search_filter(model, q, extra_fields=None, ignore_fields=None):
    if not q:
        return Q()

    ignore_fields = set(ignore_fields or [])
    fields = []
    try:
        fields = [f.name for f in model._meta.fields if not f.is_relation]
    except:
        fields = []

    if extra_fields:
        fields.extend(extra_fields)

    search_q = Q()
    for field in fields:
        if field in ignore_fields:
            continue
        search_q |= Q(**{f"{field}__icontains": q})
    return search_q


def _apply_field_filters(queryset, params, filter_fields=None):
    """
    Apply exact-match filters from query params.
    filter_fields: list of dicts like:
        { "param": "status", "field": "status" }
        { "param": "customer", "field": "customer_id" }
        { "param": "is_active", "field": "is_active", "type": "bool" }
    """
    if not filter_fields:
        return queryset
    for ff in filter_fields:
        param_name = ff["param"]
        field_name = ff.get("field", param_name)
        val = params.get(param_name)
        if val is None or str(val).strip() == "":
            continue
        val = str(val).strip()
        field_type = ff.get("type", "str")
        if field_type == "bool":
            queryset = queryset.filter(**{field_name: val.lower() in ("true", "1", "yes")})
        else:
            queryset = queryset.filter(**{field_name: val})
    return queryset


def _apply_datagrid(queryset, model, params, extra_search_fields=None, ignore_fields=None, filter_fields=None):
    pg, limit, q = _get_pg_limit_q(params)
    if q:
        queryset = queryset.filter(
            _build_search_filter(
                model=model,
                q=q,
                extra_fields=extra_search_fields,
                ignore_fields=ignore_fields,
            )
        )
    queryset = _apply_field_filters(queryset, params, filter_fields)
    count = queryset.count()
    start = pg * limit
    end = start + limit
    return queryset[start:end], count


def _row_get(row, *keys):
    for k in keys:
        if k in row and row[k] is not None and str(row[k]).strip() != "":
            return row[k]
    return None


def _parse_decimal_cell(val, default=None):
    if default is None:
        default = Decimal("0")
    if val is None or str(val).strip() == "":
        return default
    try:
        return Decimal(str(val).replace(",", "").strip())
    except Exception:
        return default


def _parse_active_cell(val):
    if val is None or str(val).strip() == "":
        return True
    s = str(val).strip().upper()
    return s not in ("NO", "FALSE", "0", "N", "INACTIVE")


class FirmService:
    @staticmethod
    def create_firm(data):
        serializer = FirmSerializer(data=data)
        if serializer.is_valid():
            firm = serializer.save()
            return BaseResponse(
                message="Firm created successfully",
                data=serializer.data,
                status=201
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )

    @staticmethod
    def get_firm(slug):
        try:
            firm = Firm.objects.get(slug=slug)
            serializer = FirmSerializer(firm)
            return BaseResponse(
                data=serializer.data,
                status=200
            )
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )

    @staticmethod
    def update_firm(slug, data):
        try:
            firm = Firm.objects.get(slug=slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404,
            )

        serializer = FirmSerializer(firm, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return BaseResponse(
                message="Firm updated successfully",
                data=serializer.data,
                status=200,
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400,
        )

    @staticmethod
    def list_firms(user=None, params=None):
        # Admins can see all firms; others only see firms they belong to
        if user and getattr(user, "user_type", None) != "ADMIN":
            firms = Firm.objects.filter(firm_users__user=user).distinct()
        else:
            firms = Firm.objects.all()

        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=firms.order_by("name"),
            model=Firm,
            params=params,
            extra_search_fields=None,
            ignore_fields=["id"],
        )
        serializer = FirmSerializer(page_qs, many=True)
        data = {"rows": serializer.data, "count": count}
        return BaseResponse(
            data=data,
            status=200
        )

    @staticmethod
    def add_user_to_firm(slug, data):
        try:
            firm = Firm.objects.get(slug=slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        # Inject firm ID
        data["firm"] = firm.id
        # Ensure user_type is FIRM_USER if not provided
        if "user_type" not in data:
             # We need to import UserTypeChoices. 
             # To avoid circular imports, maybe just string "FIRM_USER" if logic permits, 
             # but better to import from accounts.choices
             from accounts.choices import UserTypeChoices
             data["user_type"] = UserTypeChoices.FIRM_USER
        
        from accounts.serializers import UserCreateSerializer
        serializer = UserCreateSerializer(data=data)
        if serializer.is_valid():
            user = serializer.save()
            return BaseResponse(
                message="User created successfully",
                data={"id": user.id},
                status=201
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )

class ProductService:
    @staticmethod
    def create_product(firm_slug, data):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
             return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        serializer = ProductSerializer(data=data)
        if serializer.is_valid():
            serializer.save(firm=firm)
            return BaseResponse(
                message="Product created successfully",
                data=serializer.data,
                status=201
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )

    @staticmethod
    def list_products(firm_slug, params=None):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
             return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        products = Product.objects.filter(firm=firm).select_related("scheme_free_product").order_by("-created_on")
        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=products,
            model=Product,
            params=params,
            extra_search_fields=None,
            ignore_fields=["id", "firm"],
            filter_fields=[
                {"param": "category", "field": "category"},
                {"param": "is_active", "field": "is_active", "type": "bool"},
            ],
        )
        serializer = ProductSerializer(page_qs, many=True)
        data = {"rows": serializer.data, "count": count}
        return BaseResponse(
            data=data,
            status=200
        )

    @staticmethod
    def bulk_import_products(firm_slug, files):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
             return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        file_obj = files.get('file')
        if not file_obj:
            return BaseResponse(
                success=False,
                message="No file uploaded",
                status=400
            )

        filename = file_obj.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            return BaseResponse(
                success=False,
                message="Only CSV and Excel files are supported",
                status=400
            )

        products_data = []
        try:
            if filename.endswith('.csv'):
                import csv
                import io
                decoded_file = file_obj.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                # Normalize headers: lower, strip whitespace
                if reader.fieldnames:
                    reader.fieldnames = [str(col).strip().lower() for col in reader.fieldnames]

                for row in reader:
                    products_data.append(row)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                sheet = wb.active
                
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 0:
                    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                    for row in rows[1:]:
                        if not any(row): continue # skip empty rows
                        row_dict = dict(zip(headers, row))
                        products_data.append(row_dict)

        except Exception as e:
            return BaseResponse(
                success=False,
                message=f"Error parsing file: {str(e)}",
                status=400
            )

        if not products_data:
            return BaseResponse(
                success=False,
                message="File contains no valid product data",
                status=400
            )

        success_count = 0
        errors = []

        with transaction.atomic():
            for i, row in enumerate(products_data):
                name_raw = _row_get(row, "itemname", "item name", "name", "product name")
                name = str(name_raw).strip() if name_raw else ""
                description = str(_row_get(row, "description") or "").strip()
                category = str(_row_get(row, "category") or "").strip()
                hsn_code = str(
                    _row_get(row, "hsnno", "hsn_no", "hsn_code", "hsn code", "hsn") or ""
                ).strip()

                if not name:
                    errors.append(f"Row {i+2}: item name / name is required")
                    continue

                if Product.objects.filter(firm=firm, name__iexact=name).exists():
                    errors.append(f"Row {i+2}: Product '{name}' already exists")
                    continue

                Product.objects.create(
                    firm=firm,
                    name=name,
                    description=description or None,
                    category=category or None,
                    hsn_code=hsn_code or None,
                    gst_percent=_parse_decimal_cell(
                        _row_get(row, "gstper", "gst_percent", "gst %")
                    ),
                    liters=_parse_decimal_cell(_row_get(row, "liters")),
                    pack=_parse_decimal_cell(_row_get(row, "pack")),
                    mrp=_parse_decimal_cell(_row_get(row, "mrp")),
                    purchase_rate=_parse_decimal_cell(
                        _row_get(row, "prate", "purchase_rate")
                    ),
                    purchase_rate_per_unit=_parse_decimal_cell(
                        _row_get(row, "prateur", "purchase_rate_per_unit")
                    ),
                    sale_rate=_parse_decimal_cell(_row_get(row, "srate", "sale_rate")),
                    rate_per_unit=_parse_decimal_cell(
                        _row_get(
                            row,
                            "rate/unit",
                            "rate/uni",
                            "rate_uni",
                            "rate per unit",
                            "rate_per_unit",
                        )
                    ),
                    product_discount=_parse_decimal_cell(
                        _row_get(row, "product_discount", "discount", "disc")
                    ),
                    is_active=_parse_active_cell(_row_get(row, "active", "is_active")),
                )
                success_count += 1

        if errors and success_count == 0:
             return BaseResponse(
                success=False,
                message="Failed to import any products",
                errors=errors,
                status=400
            )

        return BaseResponse(
            message=f"Successfully imported {success_count} products",
            data={"success_count": success_count, "errors": errors},
            status=201 if success_count > 0 else 400
        )


class VendorOrderService:
    @staticmethod
    def create_order(firm_slug, data):
        """Create a new vendor order with items"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        # Add firm to data
        data['firm'] = firm.id
        
        serializer = VendorOrderCreateSerializer(data=data)
        if serializer.is_valid():
            with transaction.atomic():
                order = serializer.save(firm=firm)
                # Return full order details
                response_serializer = VendorOrderSerializer(order)
                return BaseResponse(
                    message="Vendor order created successfully",
                    data=response_serializer.data,
                    status=201
                )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )
    
    @staticmethod
    def list_orders(firm_slug, filters=None, params=None):
        """List all vendor orders for a firm"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        orders = VendorOrder.objects.filter(firm=firm).order_by("-order_date", "-created_on")
        
        # Apply filters if provided
        if filters:
            if 'vendor' in filters:
                orders = orders.filter(vendor_id=filters['vendor'])
            if 'order_status' in filters:
                orders = orders.filter(order_status=filters['order_status'])
            if 'payment_status' in filters:
                orders = orders.filter(payment_status=filters['payment_status'])
        
        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=orders,
            model=VendorOrder,
            params=params,
            extra_search_fields=["vendor__vendor_name"],
            ignore_fields=["id", "firm", "vendor"],
            filter_fields=[
                {"param": "order_status", "field": "order_status"},
                {"param": "payment_status", "field": "payment_status"},
                {"param": "vendor", "field": "vendor_id"},
            ],
        )

        serializer = VendorOrderSerializer(page_qs, many=True)
        data = {"rows": serializer.data, "count": count}
        return BaseResponse(
            data=data,
            status=200
        )
    
    @staticmethod
    def get_order(firm_slug, order_id):
        """Get details of a specific vendor order"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        try:
            order = VendorOrder.objects.get(id=order_id, firm=firm)
            serializer = VendorOrderSerializer(order)
            return BaseResponse(
                data=serializer.data,
                status=200
            )
        except VendorOrder.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Order not found",
                status=404
            )
    
    @staticmethod
    def update_order(firm_slug, order_id, data):
        """Update a vendor order"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        try:
            order = VendorOrder.objects.get(id=order_id, firm=firm)
        except VendorOrder.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Order not found",
                status=404
            )
        
        serializer = VendorOrderCreateSerializer(order, data=data, partial=True)
        if serializer.is_valid():
            with transaction.atomic():
                order = serializer.save()
                response_serializer = VendorOrderSerializer(order)
                return BaseResponse(
                    message="Order updated successfully",
                    data=response_serializer.data,
                    status=200
                )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )
    
    @staticmethod
    def delete_order(firm_slug, order_id):
        """Delete a vendor order"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        try:
            order = VendorOrder.objects.get(id=order_id, firm=firm)
            order.delete()
            return BaseResponse(
                message="Order deleted successfully",
                status=200
            )
        except VendorOrder.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Order not found",
                status=404
            )
    
    @staticmethod
    def receive_order(firm_slug, order_id, data, user=None):
        """Mark order as received and create product batches"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Firm not found",
                status=404
            )
        
        try:
            order = VendorOrder.objects.get(id=order_id, firm=firm)
        except VendorOrder.DoesNotExist:
            return BaseResponse(
                success=False,
                message="Order not found",
                status=404
            )
        
        if order.order_status == 'RECEIVED' or order.order_status == 'COMPLETED':
            return BaseResponse(
                success=False,
                message="Order has already been received",
                status=400
            )
        
        with transaction.atomic():
            # Process received quantities if provided
            items_data = data.get('items', [])
            for item_data in items_data:
                try:
                    order_item = VendorOrderItem.objects.get(id=item_data['id'], order=order)
                    order_item.quantity_received = item_data.get('quantity_received', order_item.quantity_ordered)
                    order_item.save()
                except (VendorOrderItem.DoesNotExist, KeyError):
                    continue

            # Update order status
            order.order_status = 'RECEIVED'
            order.received_date = timezone.now()
            order.save()
            
            # Create product batches for all items
            batches_created = []
            for item in order.items.all():
                batch = item.create_product_batch()
                if batch:
                    batches_created.append(batch.id)
                    qty = item.quantity_received or 0
                    if qty > 0:
                        StockLedgerEntry.objects.create(
                            firm=firm,
                            product=item.product,
                            product_batch=batch,
                            quantity_delta=qty,
                            entry_type=StockLedgerEntryType.VENDOR_RECEIPT,
                            vendor_order_item=item,
                            created_by=user,
                            note=f"Vendor order {order.order_number}",
                        )

            response_serializer = VendorOrderSerializer(order)
            return BaseResponse(
                message=f"Order received successfully. {len(batches_created)} product batch(es) created.",
                data={
                    "order": response_serializer.data,
                    "batches_created": batches_created
                },
                status=200
            )

    @staticmethod
    def bulk_import_orders(firm_slug, files):
        """Bulk import vendor orders from CSV/XLSX file"""
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        file_obj = files.get('file')
        if not file_obj:
            return BaseResponse(success=False, message="No file uploaded", status=400)

        filename = file_obj.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            return BaseResponse(success=False, message="Only CSV and Excel files are supported", status=400)

        # Parse file into flat rows
        rows_data = []
        try:
            if filename.endswith('.csv'):
                import csv, io
                decoded_file = file_obj.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                if reader.fieldnames:
                    reader.fieldnames = [str(col).strip().lower() for col in reader.fieldnames]
                for row in reader:
                    rows_data.append(row)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                sheet = wb.active
                all_rows = list(sheet.iter_rows(values_only=True))
                if len(all_rows) > 0:
                    headers = [str(h).strip().lower() if h else '' for h in all_rows[0]]
                    for row in all_rows[1:]:
                        if not any(row):
                            continue
                        rows_data.append(dict(zip(headers, row)))
        except Exception as e:
            return BaseResponse(success=False, message=f"Error parsing file: {str(e)}", status=400)

        if not rows_data:
            return BaseResponse(success=False, message="File contains no valid data", status=400)

        # Group rows by order_number
        from collections import defaultdict
        order_groups = defaultdict(list)
        errors = []

        for i, row in enumerate(rows_data):
            row_num = i + 2  # Excel row number (1-indexed header + 1)
            order_number = str(row.get('order_number', '')).strip() if row.get('order_number') else ''
            if not order_number:
                errors.append(f"Row {row_num}: 'order_number' is required")
                continue
            order_groups[order_number].append((row_num, row))

        orders_created = 0

        with transaction.atomic():
            for order_number, group_rows in order_groups.items():
                first_row_num, first_row = group_rows[0]

                # --- Resolve Vendor ---
                vendor_name = str(first_row.get('vendor_name', first_row.get('vendor', ''))).strip() if first_row.get('vendor_name', first_row.get('vendor')) else ''
                if not vendor_name:
                    errors.append(f"Row {first_row_num} (Order {order_number}): 'vendor_name' is required")
                    continue

                vendor = Vendor.objects.filter(firm=firm, vendor_name__iexact=vendor_name).first()
                if not vendor:
                    errors.append(f"Row {first_row_num} (Order {order_number}): Vendor '{vendor_name}' not found")
                    continue

                # --- Parse order-level fields ---
                order_date_raw = first_row.get('order_date', '')
                if not order_date_raw:
                    errors.append(f"Row {first_row_num} (Order {order_number}): 'order_date' is required")
                    continue

                from django.utils.dateparse import parse_datetime, parse_date
                import datetime
                order_date = None
                if isinstance(order_date_raw, datetime.datetime):
                    order_date = order_date_raw
                elif isinstance(order_date_raw, datetime.date):
                    order_date = datetime.datetime.combine(order_date_raw, datetime.time.min)
                else:
                    order_date_raw_str = str(order_date_raw).strip()
                    order_date = parse_datetime(order_date_raw_str)
                    if not order_date:
                        parsed_d = parse_date(order_date_raw_str)
                        if parsed_d:
                            order_date = datetime.datetime.combine(parsed_d, datetime.time.min)

                if not order_date:
                    errors.append(f"Row {first_row_num} (Order {order_number}): Invalid 'order_date' format")
                    continue

                # Check if order_number already exists
                if VendorOrder.objects.filter(order_number=order_number).exists():
                    errors.append(f"Row {first_row_num}: Order '{order_number}' already exists")
                    continue

                vendor_invoice_number = str(first_row.get('vendor_invoice_number', '')).strip() if first_row.get('vendor_invoice_number') else ''
                notes = str(first_row.get('notes', '')).strip() if first_row.get('notes') else ''

                # --- Create VendorOrder ---
                order = VendorOrder.objects.create(
                    firm=firm,
                    vendor=vendor,
                    order_number=order_number,
                    order_date=order_date,
                    vendor_invoice_number=vendor_invoice_number,
                    notes=notes,
                    total_amount=0
                )

                # --- Create VendorOrderItems ---
                total_amount = 0
                items_created = 0

                for row_num, row in group_rows:
                    product_name = str(row.get('product_name', row.get('product', ''))).strip() if row.get('product_name', row.get('product')) else ''
                    if not product_name:
                        errors.append(f"Row {row_num} (Order {order_number}): 'product_name' is required")
                        continue

                    product = Product.objects.filter(firm=firm, name__iexact=product_name).first()
                    if not product:
                        errors.append(f"Row {row_num} (Order {order_number}): Product '{product_name}' not found")
                        continue

                    try:
                        qty = int(float(str(row.get('quantity_ordered', 0))))
                        cost_price = float(str(row.get('cost_price_per_unit', 0)))
                        ss_price = float(str(row.get('selling_price_super_seller', 0)))
                        dist_price = float(str(row.get('selling_price_distributor', 0)))
                    except (ValueError, TypeError) as e:
                        errors.append(f"Row {row_num} (Order {order_number}): Invalid numeric value - {str(e)}")
                        continue

                    batch_number = str(row.get('batch_number', '')).strip() if row.get('batch_number') else ''
                    if not batch_number:
                        errors.append(f"Row {row_num} (Order {order_number}): 'batch_number' is required")
                        continue

                    # Parse optional dates
                    mfg_date = None
                    exp_date = None
                    mfg_raw = row.get('manufacturing_date')
                    exp_raw = row.get('expiry_date')

                    if mfg_raw:
                        if isinstance(mfg_raw, (datetime.datetime, datetime.date)):
                            mfg_date = mfg_raw if isinstance(mfg_raw, datetime.datetime) else datetime.datetime.combine(mfg_raw, datetime.time.min)
                        else:
                            mfg_date = parse_datetime(str(mfg_raw).strip()) or (parse_date(str(mfg_raw).strip()) and datetime.datetime.combine(parse_date(str(mfg_raw).strip()), datetime.time.min))

                    if exp_raw:
                        if isinstance(exp_raw, (datetime.datetime, datetime.date)):
                            exp_date = exp_raw if isinstance(exp_raw, datetime.datetime) else datetime.datetime.combine(exp_raw, datetime.time.min)
                        else:
                            exp_date = parse_datetime(str(exp_raw).strip()) or (parse_date(str(exp_raw).strip()) and datetime.datetime.combine(parse_date(str(exp_raw).strip()), datetime.time.min))

                    VendorOrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity_ordered=qty,
                        cost_price_per_unit=cost_price,
                        selling_price_super_seller=ss_price,
                        selling_price_distributor=dist_price,
                        batch_number=batch_number,
                        manufacturing_date=mfg_date,
                        expiry_date=exp_date
                    )
                    total_amount += qty * cost_price
                    items_created += 1

                if items_created == 0:
                    # No items created for this order, delete the empty order
                    order.delete()
                    errors.append(f"Order '{order_number}': No valid items, order was not created")
                else:
                    order.total_amount = total_amount
                    order.save()
                    orders_created += 1

        if errors and orders_created == 0:
            return BaseResponse(
                success=False,
                message="Failed to import any orders",
                errors=errors,
                status=400
            )

        return BaseResponse(
            message=f"Successfully imported {orders_created} vendor order(s)",
            data={"orders_created": orders_created, "errors": errors},
            status=201 if orders_created > 0 else 400
        )

class FirmUserService:
    @staticmethod
    def list_firm_users(firm_slug, params=None):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        # Exclude admin and owner users from firm user management
        firm_users = FirmUsers.objects.filter(
            firm=firm,
        ).exclude(user__user_type='ADMIN').select_related("user", "firm").order_by("-created_on")

        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=firm_users,
            model=FirmUsers,
            params=params,
            extra_search_fields=[
                "user__email",
                "user__full_name",
                "user__phone",
                "user__username",
                "firm__name",
            ],
            ignore_fields=["id", "user", "firm"],
            filter_fields=[
                {"param": "role", "field": "role"},
            ],
        )

        serializer = FirmUserSerializer(page_qs, many=True)
        data = {"rows": serializer.data, "count": count}
        return BaseResponse(data=data, status=200)

    @staticmethod
    def create_firm_user(firm_slug, data):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = FirmUserCreateSerializer(data=data, context={'firm': firm})
        if serializer.is_valid():
            firm_user = serializer.save()
            return BaseResponse(
                message="User created and added to firm successfully",
                data=FirmUserSerializer(firm_user).data,
                status=201
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )

    @staticmethod
    def get_firm_user(firm_slug, user_id):
        try:
            firm_user = FirmUsers.objects.get(id=user_id, firm__slug=firm_slug)
            serializer = FirmUserSerializer(firm_user)
            return BaseResponse(data=serializer.data, status=200)
        except FirmUsers.DoesNotExist:
            return BaseResponse(success=False, message="User not found", status=404)

    @staticmethod
    def update_firm_user(firm_slug, user_id, data):
        try:
            firm_user = FirmUsers.objects.get(id=user_id, firm__slug=firm_slug)
        except FirmUsers.DoesNotExist:
            return BaseResponse(success=False, message="User not found", status=404)

        serializer = FirmUserUpdateSerializer(firm_user, data=data, partial=True)
        if serializer.is_valid():
            firm_user = serializer.save()
            return BaseResponse(
                message="User updated successfully",
                data=FirmUserSerializer(firm_user).data,
                status=200
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400
        )

    @staticmethod
    def delete_firm_user(firm_slug, user_id):
        # We implementation deactivation instead of hard delete for users
        try:
            firm_user = FirmUsers.objects.get(id=user_id, firm__slug=firm_slug)
            user = firm_user.user
            user.is_active = False
            user.save()
            return BaseResponse(message="User deactivated successfully", status=200)
        except FirmUsers.DoesNotExist:
            return BaseResponse(success=False, message="User not found", status=404)


class VendorService:
    @staticmethod
    def list_vendors(firm_slug, params=None):
        try:
            vendors = Vendor.objects.filter(firm__slug=firm_slug).order_by("-created_on")
            params = params or {}
            page_qs, count = _apply_datagrid(
                queryset=vendors,
                model=Vendor,
                params=params,
                extra_search_fields=None,
                ignore_fields=["id", "firm"],
            )
            serializer = VendorSerializer(page_qs, many=True)
            data = {"rows": serializer.data, "count": count}
            return BaseResponse(data=data, status=200)
        except Exception as e:
            return BaseResponse(success=False, message=str(e), status=500)

    @staticmethod
    def create_vendor(firm_slug, data):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = VendorSerializer(data=data)
        if serializer.is_valid():
            serializer.save(firm=firm)
            return BaseResponse(message="Vendor created successfully", data=serializer.data, status=201)
        return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

    @staticmethod
    def get_vendor(firm_slug, vendor_id):
        try:
            vendor = Vendor.objects.get(id=vendor_id, firm__slug=firm_slug)
            serializer = VendorSerializer(vendor)
            return BaseResponse(data=serializer.data, status=200)
        except Vendor.DoesNotExist:
            return BaseResponse(success=False, message="Vendor not found", status=404)

    @staticmethod
    def update_vendor(firm_slug, vendor_id, data):
        try:
            vendor = Vendor.objects.get(id=vendor_id, firm__slug=firm_slug)
        except Vendor.DoesNotExist:
            return BaseResponse(success=False, message="Vendor not found", status=404)

        serializer = VendorSerializer(vendor, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return BaseResponse(message="Vendor updated successfully", data=serializer.data, status=200)
        return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

    @staticmethod
    def delete_vendor(firm_slug, vendor_id):
        try:
            vendor = Vendor.objects.get(id=vendor_id, firm__slug=firm_slug)
            vendor.delete()
            return BaseResponse(message="Vendor deleted successfully", status=200)
        except Vendor.DoesNotExist:
            return BaseResponse(success=False, message="Vendor not found", status=404)


class CustomerService:
    @staticmethod
    def list_customers(firm_slug, params=None):
        try:
            customers = Customer.objects.filter(firm__slug=firm_slug).order_by("-created_on")
            params = params or {}
            page_qs, count = _apply_datagrid(
                queryset=customers,
                model=Customer,
                params=params,
                extra_search_fields=None,
                ignore_fields=["id", "firm"],
                filter_fields=[
                    {"param": "customer_type", "field": "customer_type"},
                    {"param": "is_active", "field": "is_active", "type": "bool"},
                ],
            )
            serializer = CustomerSerializer(page_qs, many=True)
            data = {"rows": serializer.data, "count": count}
            return BaseResponse(data=data, status=200)
        except Exception as e:
            return BaseResponse(success=False, message=str(e), status=500)

    @staticmethod
    def create_customer(firm_slug, data):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = CustomerSerializer(data=data)
        if serializer.is_valid():
            serializer.save(firm=firm)
            return BaseResponse(message="Customer created successfully", data=serializer.data, status=201)
        return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

    @staticmethod
    def get_customer(firm_slug, customer_id):
        try:
            customer = Customer.objects.get(id=customer_id, firm__slug=firm_slug)
            serializer = CustomerSerializer(customer)
            return BaseResponse(data=serializer.data, status=200)
        except Customer.DoesNotExist:
            return BaseResponse(success=False, message="Customer not found", status=404)

    @staticmethod
    def update_customer(firm_slug, customer_id, data):
        try:
            customer = Customer.objects.get(id=customer_id, firm__slug=firm_slug)
        except Customer.DoesNotExist:
            return BaseResponse(success=False, message="Customer not found", status=404)

        serializer = CustomerSerializer(customer, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return BaseResponse(message="Customer updated successfully", data=serializer.data, status=200)
        return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

    @staticmethod
    def delete_customer(firm_slug, customer_id):
        try:
            customer = Customer.objects.get(id=customer_id, firm__slug=firm_slug)
            customer.delete()
            return BaseResponse(message="Customer deleted successfully", status=200)
        except Customer.DoesNotExist:
            return BaseResponse(success=False, message="Customer not found", status=404)

    @staticmethod
    def list_fssai_expiry_alerts(firm_slug, params=None):
        """Retailers with FSSAI expired or expiring within the next 7 days (requires fssai_expiry set)."""
        from datetime import timedelta

        try:
            now = timezone.now()
            cutoff = now + timedelta(days=7)
            qs = Customer.objects.filter(
                firm__slug=firm_slug,
                fssai_expiry__isnull=False,
                fssai_expiry__lte=cutoff,
            )
            params = params or {}
            window = (params.get("window") or "").strip().lower()
            if window == "expired":
                qs = qs.filter(fssai_expiry__lt=now)
            elif window == "expiring":
                qs = qs.filter(fssai_expiry__gte=now, fssai_expiry__lte=cutoff)

            qs = qs.order_by("fssai_expiry")
            page_qs, count = _apply_datagrid(
                queryset=qs,
                model=Customer,
                params=params,
                extra_search_fields=["reference_code"],
                ignore_fields=["id", "firm", "fssai_document"],
                filter_fields=None,
            )
            serializer = CustomerSerializer(page_qs, many=True)
            data = {"rows": serializer.data, "count": count}
            return BaseResponse(data=data, status=200)
        except Exception as e:
            return BaseResponse(success=False, message=str(e), status=500)


class ProductCrudService:
    @staticmethod
    def get_product(firm_slug, product_id):
        try:
            product = Product.objects.select_related("scheme_free_product").get(id=product_id, firm__slug=firm_slug)
            serializer = ProductSerializer(product)
            return BaseResponse(data=serializer.data, status=200)
        except Product.DoesNotExist:
            return BaseResponse(success=False, message="Product not found", status=404)

    @staticmethod
    def update_product(firm_slug, product_id, data):
        try:
            product = Product.objects.get(id=product_id, firm__slug=firm_slug)
        except Product.DoesNotExist:
            return BaseResponse(success=False, message="Product not found", status=404)

        serializer = ProductSerializer(product, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return BaseResponse(message="Product updated successfully", data=serializer.data, status=200)
        return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

    @staticmethod
    def delete_product(firm_slug, product_id):
        try:
            product = Product.objects.get(id=product_id, firm__slug=firm_slug)
            product.delete()
            return BaseResponse(message="Product deleted successfully", status=200)
        except Product.DoesNotExist:
            return BaseResponse(success=False, message="Product not found", status=404)


class RetailerOrderService:
    """Salesman orders to retailers; listed by firm; invoiced by admin via invoice API."""

    @staticmethod
    def list_retailer_orders(firm_slug, user, params=None):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        qs = (
            RetailerOrder.objects.filter(firm=firm)
            .select_related("customer", "created_by")
            .order_by("-created_on")
        )

        if user.user_type == 'FIRM_USER':
            try:
                firm_user = FirmUsers.objects.get(user=user, firm=firm)
                if firm_user.role not in ['ADMIN', 'FIRM_ADMIN']:
                    qs = qs.filter(created_by=user)
            except FirmUsers.DoesNotExist:
                qs = qs.filter(created_by=user)
        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=qs,
            model=RetailerOrder,
            params=params,
            extra_search_fields=[
                "customer__business_name",
                "reference",
                "notes",
            ],
            ignore_fields=["id", "firm", "customer", "created_by"],
            filter_fields=[
                {"param": "status", "field": "status"},
                {"param": "customer", "field": "customer_id"},
            ],
        )
        data = {
            "rows": RetailerOrderSerializer(page_qs, many=True).data,
            "count": count,
        }
        return BaseResponse(data=data, status=200)

    @staticmethod
    def create_retailer_order(firm_slug, data, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = RetailerOrderCreateSerializer(
            data=data,
            context={"firm": firm, "request_user": user},
        )
        if serializer.is_valid():
            order = serializer.save()
            return BaseResponse(
                message="Retailer order created",
                data=RetailerOrderSerializer(order).data,
                status=201,
            )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400,
        )

    @staticmethod
    def get_retailer_order(firm_slug, order_id, user=None):
        try:
            order = RetailerOrder.objects.get(id=order_id, firm__slug=firm_slug)
        except RetailerOrder.DoesNotExist:
            return BaseResponse(success=False, message="Order not found", status=404)

        if user is not None and getattr(user, "user_type", None) == "FIRM_USER":
            try:
                firm_user = FirmUsers.objects.get(user=user, firm=order.firm)
                if firm_user.role not in ("ADMIN", "FIRM_ADMIN"):
                    if str(order.created_by_id) != str(user.id):
                        return BaseResponse(
                            success=False,
                            message="Forbidden",
                            status=403,
                        )
            except FirmUsers.DoesNotExist:
                if str(order.created_by_id) != str(user.id):
                    return BaseResponse(
                        success=False,
                        message="Forbidden",
                        status=403,
                    )

        return BaseResponse(
            data=RetailerOrderSerializer(order).data,
            status=200,
        )


class InvoiceService:
    """Terminal / finished states where requesting changes is not allowed."""
    REQUEST_CHANGES_BLOCKED_STATUSES = frozenset({"CLOSED", "CANCELLED", "REJECTED"})

    @staticmethod
    def auto_close_delivered_invoices(firm=None):
        """
        Set status to CLOSED when delivered_at is at least 2 days ago and the invoice
        is still in a post-delivery state (not yet closed).
        """
        from datetime import timedelta

        threshold = timezone.now() - timedelta(days=2)
        qs = Invoice.objects.filter(
            delivered_at__isnull=False,
            delivered_at__lte=threshold,
            status__in=["DELIVERED", "PARTIALLY_PAID", "PAID"],
        )
        if firm is not None:
            qs = qs.filter(firm=firm)
        return qs.update(status="CLOSED", updated_on=timezone.now())

    @staticmethod
    def list_invoices(firm_slug, user, params=None):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        InvoiceService.auto_close_delivered_invoices(firm=firm)

        invoices = Invoice.objects.filter(firm=firm).select_related("customer", "created_by", "approved_by").order_by("-created_on")
        
        # If the user is just a regular firm user, they can only see invoices they created
        # unless we want them to see all of them. Assuming regular users only see their own.
        if user.user_type == 'FIRM_USER':
            # Check if they are ADMIN or FIRM_ADMIN role
            try:
                firm_user = FirmUsers.objects.get(user=user, firm=firm)
                if firm_user.role not in ['ADMIN', 'FIRM_ADMIN']:
                    invoices = invoices.filter(created_by=user)
            except FirmUsers.DoesNotExist:
                invoices = invoices.filter(created_by=user)

        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=invoices,
            model=Invoice,
            params=params,
            extra_search_fields=[
                "customer__business_name",
                "invoice_number",
            ],
            ignore_fields=["id", "firm", "customer", "created_by", "approved_by"],
            filter_fields=[
                {"param": "status", "field": "status"},
                {"param": "customer", "field": "customer_id"},
                {"param": "is_printed", "field": "is_printed", "type": "bool"},
            ],
        )

        serializer = InvoiceSerializer(page_qs, many=True)
        data = {"rows": serializer.data, "count": count}
        return BaseResponse(data=data, status=200)

    @staticmethod
    def create_invoice(firm_slug, data, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = InvoiceFromRetailerOrdersSerializer(
            data=data,
            context={"firm": firm, "request_user": user},
        )
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    invoice = serializer.save()
                    response_serializer = InvoiceSerializer(invoice)
                    return BaseResponse(
                        message="Invoice created from retailer orders",
                        data=response_serializer.data,
                        status=201,
                    )
            except DRFValidationError as ex:
                return BaseResponse(
                    success=False,
                    message=_drf_validation_detail_message(ex),
                    status=400,
                )
        return BaseResponse(
            success=False,
            message="Invalid data",
            errors=serializer.errors,
            status=400,
        )

    @staticmethod
    def get_invoice(firm_slug, invoice_id):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        InvoiceService.auto_close_delivered_invoices(firm=firm)

        try:
            invoice = (
                Invoice.objects.select_related("customer", "created_by", "approved_by")
                .prefetch_related("items__product", "items__product_batch", "source_orders")
                .get(id=invoice_id, firm=firm)
            )
            serializer = InvoiceSerializer(invoice)
            return BaseResponse(data=serializer.data, status=200)
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)

    INVOICE_EDITABLE_STATUSES = frozenset({"PENDING_APPROVAL", "CHANGES_REQUESTED"})

    @staticmethod
    def update_invoice(firm_slug, invoice_id, data, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = InvoiceUpdateSerializer(data=data, context={"firm": firm})
        if not serializer.is_valid():
            return BaseResponse(
                success=False,
                message="Invalid data",
                errors=serializer.errors,
                status=400,
            )

        try:
            invoice = Invoice.objects.select_related("customer").get(
                id=invoice_id, firm=firm
            )
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)

        if invoice.status not in InvoiceService.INVOICE_EDITABLE_STATUSES:
            return BaseResponse(
                success=False,
                message=(
                    "Invoice lines can only be edited while pending approval "
                    "or after changes are requested."
                ),
                status=400,
            )

        if Payment.objects.filter(invoice_id=invoice_id).exists():
            return BaseResponse(
                success=False,
                message="Cannot edit an invoice that has payments recorded.",
                status=400,
            )

        overrides = serializer.validated_data["line_items"]
        lines_to_invoice = build_lines_from_line_overrides(
            overrides, invoice.customer
        )

        try:
            with transaction.atomic():
                invoice_locked = Invoice.objects.select_related("customer").select_for_update().get(
                    id=invoice_id, firm=firm
                )
                reverse_invoice_line_allocations(firm, user, invoice_locked)
                apply_invoice_lines_to_invoice(
                    invoice_locked,
                    firm,
                    user,
                    lines_to_invoice,
                    "Invoice revision",
                )
        except DRFValidationError as ex:
            return BaseResponse(
                success=False,
                message=_drf_validation_detail_message(ex),
                status=400,
            )

        invoice = (
            Invoice.objects.select_related("customer", "created_by", "approved_by")
            .prefetch_related("items__product", "items__product_batch", "source_orders")
            .get(id=invoice_id, firm=firm)
        )
        response_serializer = InvoiceSerializer(invoice)
        return BaseResponse(
            message="Invoice updated successfully",
            data=response_serializer.data,
            status=200,
        )
        
    @staticmethod
    def _generate_invoice_number(firm):
        # Format: FIRMCODE-INV-00001
        last_invoice = Invoice.objects.filter(firm=firm, invoice_number__isnull=False).order_by('invoice_number').last()
        if last_invoice and last_invoice.invoice_number:
            try:
                prefix, num_str = last_invoice.invoice_number.rsplit('-', 1)
                num = int(num_str) + 1
            except ValueError:
                num = 1
        else:
            num = 1
        
        return f"{firm.code.upper()}-INV-{num:05d}"

    @staticmethod
    def approve_invoice(firm_slug, invoice_id, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        # try:
        #     invoice = Invoice.objects.select_for_update().get(id=invoice_id, firm=firm)
        # except Invoice.DoesNotExist:
        #     return BaseResponse(success=False, message="Invoice not found", status=404)

        # if invoice.status == 'APPROVED':
        #     return BaseResponse(success=False, message="Invoice is already approved", status=400)

        with transaction.atomic():  # ✅ Start transaction FIRST

            try:
                invoice = Invoice.objects.select_for_update().get(id=invoice_id, firm=firm)
            except Invoice.DoesNotExist:
                return BaseResponse(success=False, message="Invoice not found", status=404)
            
            if invoice.status == 'APPROVED':
                return BaseResponse(success=False, message="Invoice is already approved", status=400)

            invoice.status = 'APPROVED'
            invoice.approved_by = user
            invoice.rejection_note = None

            # ⚠️ Important: invoice number generation should also be inside txn
            invoice.invoice_number = InvoiceService._generate_invoice_number(firm)

            invoice.save()

            response_serializer = InvoiceSerializer(invoice)

            return BaseResponse(
                message="Invoice approved successfully",
                data=response_serializer.data,
                status=200
            )

    @staticmethod
    def request_changes(firm_slug, invoice_id, data):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        try:
            invoice = Invoice.objects.get(id=invoice_id, firm=firm)
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)

        if invoice.status in InvoiceService.REQUEST_CHANGES_BLOCKED_STATUSES:
            return BaseResponse(
                success=False,
                message="Changes cannot be requested for closed, cancelled, or rejected invoices.",
                status=400,
            )

        note = data.get('note', '')
        if not note:
            return BaseResponse(success=False, message="Change request must include a note", status=400)

        invoice.status = 'CHANGES_REQUESTED'
        invoice.rejection_note = note
        invoice.save()

        response_serializer = InvoiceSerializer(invoice)
        return BaseResponse(
            message="Changes requested successfully",
            data=response_serializer.data,
            status=200
        )

    VALID_STATUS_TRANSITIONS = {
        "APPROVED": ["OUT_FOR_DELIVERY", "CANCELLED"],
        "OUT_FOR_DELIVERY": ["DELIVERED", "CANCELLED"],
        # Post-delivery: close or cancel only; partial/full payment uses Payment rows + payment_status.
        "DELIVERED": ["CLOSED", "CANCELLED"],
        "PARTIALLY_PAID": ["CLOSED", "CANCELLED"],
        "PAID": ["CLOSED"],
        "CHANGES_REQUESTED": ["PENDING_APPROVAL", "CANCELLED"],
        "PENDING_APPROVAL": ["CANCELLED"],
    }

    @staticmethod
    def print_invoice(firm_slug, invoice_id, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        try:
            invoice = Invoice.objects.select_related(
                "customer", "created_by", "approved_by"
            ).prefetch_related(
                "items__product", "items__product_batch"
            ).get(id=invoice_id, firm=firm)
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)

        if not invoice.is_printed:
            invoice.is_printed = True
            invoice.printed_on = timezone.now()
            invoice.printed_by = user
            invoice.save(update_fields=["is_printed", "printed_on", "printed_by", "updated_on"])

        serializer = InvoiceSerializer(invoice)
        return BaseResponse(
            message="Invoice marked as printed",
            data=serializer.data,
            status=200,
        )

    @staticmethod
    def batch_print_invoices(firm_slug, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        invoices = Invoice.objects.filter(
            firm=firm, status="APPROVED", is_printed=False,
        ).select_related(
            "customer", "created_by", "approved_by"
        ).prefetch_related(
            "items__product", "items__product_batch"
        ).order_by("created_on")

        if not invoices.exists():
            return BaseResponse(
                success=False,
                message="No approved unprinted invoices found",
                status=404,
            )

        now = timezone.now()
        ids = list(invoices.values_list("id", flat=True))
        Invoice.objects.filter(id__in=ids).update(
            is_printed=True, printed_on=now, printed_by=user,
        )
        updated_invoices = Invoice.objects.filter(id__in=ids).select_related(
            "customer", "created_by", "approved_by"
        ).prefetch_related("items__product", "items__product_batch")
        serializer = InvoiceSerializer(updated_invoices, many=True)
        return BaseResponse(
            message=f"{len(ids)} invoice(s) marked as printed",
            data={
                "rows": serializer.data,
                "count": len(ids),
                "printed_count": len(ids),
                "printed_ids": [str(i) for i in ids],
            },
            status=200,
        )

    @staticmethod
    def update_invoice_status(firm_slug, invoice_id, data, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        new_status = data.get("status", "").strip()
        if not new_status:
            return BaseResponse(success=False, message="'status' is required", status=400)

        valid_statuses = [c[0] for c in Invoice._meta.get_field("status").choices]
        if new_status not in valid_statuses:
            return BaseResponse(
                success=False,
                message=f"Invalid status '{new_status}'",
                status=400,
            )

        with transaction.atomic():
            try:
                invoice = Invoice.objects.select_for_update().get(id=invoice_id, firm=firm)
            except Invoice.DoesNotExist:
                return BaseResponse(success=False, message="Invoice not found", status=404)

            allowed = InvoiceService.VALID_STATUS_TRANSITIONS.get(invoice.status, [])
            if new_status not in allowed:
                return BaseResponse(
                    success=False,
                    message=f"Cannot transition from '{invoice.status}' to '{new_status}'. Allowed: {allowed}",
                    status=400,
                )

            invoice.status = new_status
            update_fields = ["status", "updated_on"]
            if new_status == "DELIVERED":
                invoice.delivered_at = timezone.now()
                update_fields.append("delivered_at")
            invoice.save(update_fields=update_fields)

        serializer = InvoiceSerializer(invoice)
        return BaseResponse(
            message=f"Invoice status updated to {new_status}",
            data=serializer.data,
            status=200,
        )

    @staticmethod
    def preview_pricing(firm_slug, data):
        """
        Simulate FEFO batch allocation and return the estimated cost breakdown
        without actually creating an invoice or modifying any inventory.
        
        Expects: { customer: <id>, items: [ { product: <id>, quantity: <int> } ] }
        Returns per-item allocation breakdown with rates and totals.
        """
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        customer_id = data.get("customer")
        items = data.get("items", [])

        if not customer_id:
            return BaseResponse(success=False, message="customer is required", status=400)

        try:
            customer = Customer.objects.get(id=customer_id, firm=firm)
        except Customer.DoesNotExist:
            return BaseResponse(success=False, message="Customer not found", status=404)

        preview_items = []
        grand_total = Decimal("0")

        for req_item in items:
            product_id = req_item.get("product")
            requested_quantity = req_item.get("quantity", 0)

            if not product_id or requested_quantity <= 0:
                continue

            try:
                product = Product.objects.get(id=product_id, firm=firm)
            except Product.DoesNotExist:
                return BaseResponse(
                    success=False, message=f"Product {product_id} not found", status=404
                )

            rate = effective_unit_rate(product, customer)
            batches = list(
                ProductBatch.objects.filter(product=product, quantity__gt=0).order_by(
                    F("expiry_date").asc(nulls_last=True), "created_on"
                )
            )

            remaining_to_allocate = int(requested_quantity)
            item_breakdown = []
            item_total = Decimal("0")

            for batch in batches:
                if remaining_to_allocate <= 0:
                    break

                qty_to_take = min(batch.quantity, remaining_to_allocate)
                remaining_to_allocate -= qty_to_take

                taxable = Decimal(qty_to_take) * rate
                item_total += taxable

                exp = batch.expiry_date
                item_breakdown.append(
                    {
                        "expiry_date": exp.isoformat() if exp else None,
                        "quantity": qty_to_take,
                        "rate": str(rate),
                        "amount": str(taxable),
                    }
                )

            if remaining_to_allocate > 0:
                return BaseResponse(
                    success=False,
                    message=(
                        f"Insufficient stock for '{product.name}'. Requested {requested_quantity}, "
                        f"available {requested_quantity - remaining_to_allocate}."
                    ),
                    status=400,
                )

            item_total = line_total_inclusive(item_total, product.gst_percent)
            grand_total += item_total
            preview_items.append(
                {
                    "product": str(product.id),
                    "product_name": product.name,
                    "requested_quantity": requested_quantity,
                    "estimated_total": str(item_total.quantize(Decimal("0.01"))),
                    "batches": item_breakdown,
                }
            )

        return BaseResponse(
            data={
                "items": preview_items,
                "grand_total": str(grand_total.quantize(Decimal("0.01"))),
                "customer_type": customer.customer_type,
            },
            status=200,
        )

class DashboardService:
    @staticmethod
    def get_firm_dashboard_data(firm_slug):
        from django.db.models import Sum, Count
        from accounts.models import FirmUsers
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        # Financial aggregates
        cash_in = Invoice.objects.filter(
            firm=firm,
            status__in=['APPROVED', 'OUT_FOR_DELIVERY', 'DELIVERED', 'PARTIALLY_PAID', 'PAID', 'CLOSED'],
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        cash_out = VendorOrder.objects.filter(firm=firm, order_status__in=['RECEIVED', 'COMPLETED']).aggregate(total=Sum('total_amount'))['total'] or 0
        profit = float(cash_in) - float(cash_out)

        # Outstanding (unpaid/partial) vendor order amounts
        outstanding = VendorOrder.objects.filter(firm=firm, payment_status__in=['UNPAID', 'PARTIAL']).aggregate(
            total=Sum('total_amount'),
            paid=Sum('amount_paid')
        )
        outstanding_amount = float(outstanding['total'] or 0) - float(outstanding['paid'] or 0)

        # Counts
        total_products = Product.objects.filter(firm=firm).count()
        total_vendors = Vendor.objects.filter(firm=firm).count()
        total_customers = Customer.objects.filter(firm=firm).count()
        total_users = FirmUsers.objects.filter(firm=firm).count()

        # Invoice stats by status
        invoice_stats = dict(
            Invoice.objects.filter(firm=firm)
            .values_list('status')
            .annotate(count=Count('id'))
            .values_list('status', 'count')
        )

        # Order stats by status
        order_stats = dict(
            VendorOrder.objects.filter(firm=firm)
            .values_list('order_status')
            .annotate(count=Count('id'))
            .values_list('order_status', 'count')
        )

        # Recent invoices
        recent_invoices = list(
            Invoice.objects.filter(firm=firm)
            .select_related('customer')
            .order_by('-created_on')[:5]
            .values('id', 'invoice_number', 'total_amount', 'status', 'customer__business_name', 'created_on')
        )
        for inv in recent_invoices:
            inv['total_amount'] = float(inv['total_amount'])
            inv['created_on'] = inv['created_on'].isoformat() if inv['created_on'] else None

        data = {
            "firm_name": firm.name,
            "cash_in": float(cash_in),
            "cash_out": float(cash_out),
            "profit": profit,
            "outstanding_amount": outstanding_amount,
            "total_products": total_products,
            "total_vendors": total_vendors,
            "total_customers": total_customers,
            "total_users": total_users,
            "invoice_stats": {
                "pending": invoice_stats.get('PENDING_APPROVAL', 0),
                "approved": invoice_stats.get('APPROVED', 0),
                "changes_requested": invoice_stats.get('CHANGES_REQUESTED', 0),
                "out_for_delivery": invoice_stats.get('OUT_FOR_DELIVERY', 0),
                "delivered": invoice_stats.get('DELIVERED', 0),
                "partially_paid": invoice_stats.get('PARTIALLY_PAID', 0),
                "paid": invoice_stats.get('PAID', 0),
                "closed": invoice_stats.get('CLOSED', 0),
                "cancelled": invoice_stats.get('CANCELLED', 0),
            },
            "order_stats": {
                "pending": order_stats.get('PENDING', 0),
                "received": order_stats.get('RECEIVED', 0),
                "completed": order_stats.get('COMPLETED', 0),
            },
            "recent_invoices": recent_invoices,
        }
        
        return BaseResponse(data=data, status=200)

    @staticmethod
    def get_admin_dashboard_data():
        from django.db.models import Sum, Count
        cash_in = Invoice.objects.filter(
            status__in=['APPROVED', 'OUT_FOR_DELIVERY', 'DELIVERED', 'PARTIALLY_PAID', 'PAID', 'CLOSED'],
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        cash_out = VendorOrder.objects.filter(order_status__in=['RECEIVED', 'COMPLETED']).aggregate(total=Sum('total_amount'))['total'] or 0
        profit = float(cash_in) - float(cash_out)

        total_firms = Firm.objects.count()
        total_invoices = Invoice.objects.count()
        total_orders = VendorOrder.objects.count()

        # Per-firm breakdown
        firms = Firm.objects.all()
        firm_breakdown = []
        for f in firms:
            f_cash_in = Invoice.objects.filter(
                firm=f,
                status__in=['APPROVED', 'OUT_FOR_DELIVERY', 'DELIVERED', 'PARTIALLY_PAID', 'PAID', 'CLOSED'],
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            f_cash_out = VendorOrder.objects.filter(firm=f, order_status__in=['RECEIVED', 'COMPLETED']).aggregate(total=Sum('total_amount'))['total'] or 0
            firm_breakdown.append({
                "name": f.name,
                "slug": f.slug,
                "cash_in": float(f_cash_in),
                "cash_out": float(f_cash_out),
                "profit": float(f_cash_in) - float(f_cash_out),
                "products": Product.objects.filter(firm=f).count(),
                "vendors": Vendor.objects.filter(firm=f).count(),
                "customers": Customer.objects.filter(firm=f).count(),
            })

        data = {
            "cash_in": float(cash_in),
            "cash_out": float(cash_out),
            "profit": profit,
            "total_firms": total_firms,
            "total_invoices": total_invoices,
            "total_orders": total_orders,
            "firm_breakdown": firm_breakdown,
        }
        
        return BaseResponse(data=data, status=200)


class PaymentService:

    @staticmethod
    def list_payments(firm_slug, invoice_id):
        try:
            invoice = Invoice.objects.get(id=invoice_id, firm__slug=firm_slug)
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)
        payments = invoice.payments.select_related("recorded_by").all()
        serializer = PaymentSerializer(payments, many=True)
        paid = payments.aggregate(s=Sum("amount"))["s"] or Decimal("0")
        pending = max(Decimal("0"), (invoice.total_amount or Decimal("0")) - paid)
        data = {
            "rows": serializer.data,
            "total_amount": str(invoice.total_amount),
            "amount_paid": str(paid.quantize(Decimal("0.01"))),
            "amount_pending": str(pending.quantize(Decimal("0.01"))),
        }
        return BaseResponse(data=data, status=200)

    @staticmethod
    def add_payment(firm_slug, invoice_id, data, user):
        try:
            invoice = Invoice.objects.get(id=invoice_id, firm__slug=firm_slug)
        except Invoice.DoesNotExist:
            return BaseResponse(success=False, message="Invoice not found", status=404)

        serializer = PaymentCreateSerializer(data=data)
        if not serializer.is_valid():
            return BaseResponse(success=False, message="Invalid data", errors=serializer.errors, status=400)

        vd = serializer.validated_data
        paid_so_far = invoice.payments.aggregate(s=Sum("amount"))["s"] or Decimal("0")
        total = invoice.total_amount or Decimal("0")
        pending = total - paid_so_far
        if vd["amount"] > pending:
            return BaseResponse(
                success=False,
                message=f"Payment amount exceeds pending balance ({pending.quantize(Decimal('0.01'))}).",
                status=400,
            )

        payment = Payment.objects.create(
            invoice=invoice,
            amount=vd["amount"],
            mode=vd["mode"],
            reference=vd.get("reference", ""),
            note=vd.get("note", ""),
            paid_on=vd["paid_on"],
            recorded_by=user,
        )
        return BaseResponse(
            message="Payment recorded",
            data=PaymentSerializer(payment).data,
            status=201,
        )

    @staticmethod
    def customer_outstanding(firm_slug, customer_id):
        try:
            customer = Customer.objects.get(id=customer_id, firm__slug=firm_slug)
        except Customer.DoesNotExist:
            return BaseResponse(success=False, message="Customer not found", status=404)

        invoices = Invoice.objects.filter(
            customer=customer,
            status__in=[
                "PENDING_APPROVAL", "APPROVED", "OUT_FOR_DELIVERY",
                "DELIVERED", "PARTIALLY_PAID",
            ],
        ).prefetch_related("payments")

        rows = []
        total_outstanding = Decimal("0")
        for inv in invoices:
            paid = inv.payments.aggregate(s=Sum("amount"))["s"] or Decimal("0")
            pending = max(Decimal("0"), (inv.total_amount or Decimal("0")) - paid)
            total_outstanding += pending
            if pending > 0:
                rows.append({
                    "invoice_id": str(inv.id),
                    "invoice_number": inv.invoice_number,
                    "total_amount": str(inv.total_amount),
                    "amount_paid": str(paid.quantize(Decimal("0.01"))),
                    "amount_pending": str(pending.quantize(Decimal("0.01"))),
                    "status": inv.status,
                    "created_on": inv.created_on.isoformat() if inv.created_on else None,
                })

        data = {
            "customer_id": str(customer.id),
            "business_name": customer.business_name,
            "total_outstanding": str(total_outstanding.quantize(Decimal("0.01"))),
            "invoices": rows,
        }
        return BaseResponse(data=data, status=200)


class StockService:
    @staticmethod
    def list_stock(firm_slug, params):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        qs = Product.objects.filter(firm=firm).order_by("name")
        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=qs,
            model=Product,
            params=params,
            extra_search_fields=["name", "product_code"],
            ignore_fields=["id", "firm", "slug", "description", "image"],
            filter_fields=[
                {"param": "is_active", "field": "is_active", "type": "bool"},
                {"param": "product", "field": "id"},
            ],
        )
        rows = []
        for p in page_qs:
            total = p.batches.aggregate(s=Sum("quantity"))["s"] or 0
            batches = [
                {
                    "id": str(b.id),
                    "expiry_date": b.expiry_date.isoformat() if b.expiry_date else None,
                    "quantity": b.quantity,
                }
                for b in p.batches.filter(quantity__gt=0).order_by(
                    F("expiry_date").asc(nulls_last=True), "created_on"
                )
            ]
            rows.append(
                {
                    "id": str(p.id),
                    "product_code": p.product_code,
                    "name": p.name,
                    "total_quantity": total,
                    "batches": batches,
                }
            )
        return BaseResponse(data={"rows": rows, "count": count}, status=200)

    @staticmethod
    def list_ledger(firm_slug, params):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        qs = StockLedgerEntry.objects.filter(firm=firm).select_related(
            "product",
            "product_batch",
            "created_by",
            "vendor_order_item__order",
            "invoice_item__invoice",
        )
        params = params or {}
        page_qs, count = _apply_datagrid(
            queryset=qs,
            model=StockLedgerEntry,
            params=params,
            extra_search_fields=["product__name", "product__product_code", "note"],
            ignore_fields=["id"],
            filter_fields=[
                {"param": "entry_type", "field": "entry_type"},
                {"param": "product", "field": "product_id"},
            ],
        )
        rows = []
        for e in page_qs:
            ref = ""
            if e.vendor_order_item_id and e.vendor_order_item:
                ref = f"Vendor order {e.vendor_order_item.order.order_number}"
            elif e.invoice_item_id and e.invoice_item:
                inv = e.invoice_item.invoice
                ref = f"Invoice {inv.invoice_number or inv.slug}"

            batch_exp = None
            if e.product_batch and e.product_batch.expiry_date:
                batch_exp = e.product_batch.expiry_date.isoformat()

            rows.append(
                {
                    "id": str(e.id),
                    "created_on": e.created_on.isoformat() if e.created_on else None,
                    "product_name": e.product.name,
                    "product_code": e.product.product_code,
                    "quantity_delta": e.quantity_delta,
                    "entry_type": e.entry_type,
                    "entry_type_display": e.get_entry_type_display(),
                    "manual_reason": e.manual_reason or "",
                    "manual_reason_display": (
                        e.get_manual_reason_display() if e.manual_reason else ""
                    ),
                    "batch_id": str(e.product_batch_id) if e.product_batch_id else None,
                    "batch_expiry": batch_exp,
                    "created_by_name": (e.created_by.full_name if e.created_by else "") or "—",
                    "note": e.note or "",
                    "reference": ref,
                }
            )
        return BaseResponse(data={"rows": rows, "count": count}, status=200)

    @staticmethod
    def manual_adjust(firm_slug, data, user):
        try:
            firm = Firm.objects.get(slug=firm_slug)
        except Firm.DoesNotExist:
            return BaseResponse(success=False, message="Firm not found", status=404)

        serializer = StockManualAdjustSerializer(data=data)
        if not serializer.is_valid():
            return BaseResponse(
                success=False, message="Invalid data", errors=serializer.errors, status=400
            )

        vd = serializer.validated_data
        product = vd["product"]
        if product.firm_id != firm.id:
            return BaseResponse(
                success=False, message="Product does not belong to this firm", status=400
            )

        direction = vd["direction"]
        qty = vd["quantity"]
        note = vd.get("note") or ""
        reason = vd["manual_reason"]

        with transaction.atomic():
            if direction == "in":
                exp = vd.get("expiry_date")
                batch, _ = ProductBatch.objects.get_or_create(
                    product=product,
                    expiry_date=exp,
                    defaults={"quantity": 0},
                )
                batch.quantity += qty
                batch.save(update_fields=["quantity", "updated_on"])
                StockLedgerEntry.objects.create(
                    firm=firm,
                    product=product,
                    product_batch=batch,
                    quantity_delta=qty,
                    entry_type=StockLedgerEntryType.MANUAL,
                    manual_reason=reason,
                    note=note,
                    created_by=user,
                )
            else:
                specified = vd.get("product_batch")
                if specified:
                    if specified.product_id != product.id:
                        return BaseResponse(
                            success=False,
                            message="Selected batch does not match the product",
                            status=400,
                        )
                    if specified.quantity < qty:
                        return BaseResponse(
                            success=False,
                            message=f"Not enough stock in batch (have {specified.quantity})",
                            status=400,
                        )
                    allocations = [(specified, qty)]
                else:
                    try:
                        allocations = allocate_batches_fefo(product, qty)
                    except DRFValidationError as ex:
                        return BaseResponse(
                            success=False,
                            message=_drf_validation_detail_message(ex),
                            status=400,
                        )

                for batch, take in allocations:
                    batch.quantity -= take
                    batch.save(update_fields=["quantity", "updated_on"])
                    StockLedgerEntry.objects.create(
                        firm=firm,
                        product=product,
                        product_batch=batch,
                        quantity_delta=-take,
                        entry_type=StockLedgerEntryType.MANUAL,
                        manual_reason=reason,
                        note=note,
                        created_by=user,
                    )

        return BaseResponse(
            message="Stock adjusted",
            data={"product": str(product.id), "direction": direction, "quantity": qty},
            status=200,
        )